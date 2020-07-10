# python3 convertToXlsx.py
# convert tasks_result.json from tasks_allcircles.js to tasks_result.xlsx
import json
import sys
import re

def get_reuslt(filename):
    with open(filename,encoding='utf-8') as fid:
        data=json.load(fid)
    return data

def convertToXlsx(data):
    from openpyxl import Workbook
    wb = Workbook()
    if type(data['title'])==type(''):
        data['title']=[data['title']]
        data['data']=[data['data']]
    for ii,title in enumerate(data['title']):
        if ii==0:
            ws = wb.active 
        else:
            ws = wb.create_sheet()
        ws.title = title
        for line in data['data'][ii]:
            ws.append(line)
    wb.save(data['outFileName'])

# def readXlsx():
#     from openpyxl import load_workbook
#     wb = load_workbook(filename = 'abc.xlsx')
#     ws = wb[wb.sheetnames[0]]
#     data = ws.values
#     return list(data)

if __name__ == "__main__":
    convertToXlsx(get_reuslt(sys.argv[1]))
