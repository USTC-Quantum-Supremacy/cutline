# [markdown]
# ## export and import  

__all__=[
    
    'getSimulation',
    'reRenderSimulationFile',

    'setFSIMTime',
    'getTask',
    'getCircuit',
    'checkState',
]
import os
import re
from openpyxl import load_workbook

# [markdown]
# ## definition  

ThisPath=os.path.split(os.path.realpath(__file__))[0]

def readXlsx(filename):
    wb = load_workbook(filename = filename)
    ws = wb[wb.sheetnames[0]]
    data = ws.values
    return list(data)

def build_info():
    class InfoClass:
        def pos(self,index):
            return self.circuitIndexes.index(index)
        def get(self,index,field):
            return self.data[self.pos(index)][self.fields.index(field)]
    Info = InfoClass()
    data=readXlsx(ThisPath+'\\circuits.xlsx')
    Info.fields=data[0]
    Info.data=data[1:]
    Info.circuitIndexes=[line[Info.fields.index('circuitIndex')] for line in Info.data]
    return Info

Info=build_info()

def fillFSIM(src):
    m=re.match(r"((?:[XY].*\n)+)",src)
    qubits={line.split(' ')[1][1:]:1 for line in m.group(1).strip().split('\n')}
    def repl(subm):
        qubits2=dict(qubits)
        output=[subm.group(1)]
        for line in subm.group(1).strip().split('\n'):
            m2=re.match(r"FSIM G(\d+)_(\d+) ",line)
            qubits2[m2.group(1)]=2
            qubits2[m2.group(2)]=2
        for k,v in qubits2.items():
            if v==1:
                output.append(f"I Q{k} {g.FSIMTime}\n")
        return ''.join(output)
    out=re.sub(r"((?:[F].*\n)+)",repl,src)
    return out


# [markdown]
# global values

class g:
    ThisPath=ThisPath
    Info=Info
    FSIMTime=1

# [markdown]
# ## simulation part

def getSimulation(index, algorithm=''):
    al='SFA'
    filename = 'circuit/'+g.Info.get(index,'name')
    if algorithm=='':
        target=g.Info.get(index,'targets')
        if re.search(r'(^|_)SFA($|_)',target):al='SFA'
        if re.search(r'(^|_)SFA_?LESSMEM($|_)',target):al='SFA_LESSMEM'
        if re.search(r'(^|_)PEPS($|_)',target):al='PEPS'
        if re.search(r'(^|_)SA($|_)',target):al='SA'
    else:
        al=algorithm
    return dict(filename=filename,algorithm=al)

def reRenderSimulationFile(index,gateArgs):
    filename = g.Info.get(index,'name')
    with open(g.ThisPath+'\\circuit\\'+filename,encoding='utf-8') as fid:
        src = fid.read()
    # todo: change arguments by `gateArgs`
    out = src 
    return out

# [markdown]
# ## experiment part

def setFSIMTime(time):
    g.FSIMTime=time

def getCircuit(index):
    filename = g.Info.get(index,'name')+'.qcis'
    with open(g.ThisPath+'\\circuit\\'+filename,encoding='utf-8') as fid:
        src = fid.read()
    out = fillFSIM(src)
    return out
def getTask(resultDir,index):
    pass
def checkState(resultDir,index):
    pass