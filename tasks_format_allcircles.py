# python3 tasks_format_allcircles.py
# convert tasks_result.json from tasks_allcircles.js to tasks_result.xlsx
import json
import sys
import re

def get_reuslt():
    with open('output/tasks_result.json',encoding='utf-8') as fid:
        data=json.load(fid)
    return data

def convertToXlsx(data):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active 
    # ws2 = wb.create_sheet()
    ws.title = data['title']
    for line in data['data']:
        ws.append(line)
    wb.save('output/tasks_result.xlsx')

# def readXlsx():
#     from openpyxl import load_workbook
#     wb = load_workbook(filename = 'abc.xlsx')
#     ws = wb[wb.sheetnames[0]]
#     data = ws.values
#     return list(data)

if __name__ == "__main__":
    convertToXlsx(get_reuslt())
